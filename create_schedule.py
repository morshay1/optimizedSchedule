import pulp
import pandas as pd
import os
import random
import copy
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from helper_functions import excel_sheets_to_items

T = range # Time slots from 8 AM to 5 PM (represented as hours)

def create_original_schedule(input_file_path, output_file_path):
    # Define sets
    C, P, E, Ce, caregiver_unavailability, patient_unavailability, patient_equipment_mapping = \
        excel_sheets_to_items(input_file_path)

    # Binary variables xcpt, where xcpt = 1 if caregiver c cares for patient p at time t using equipment e
    xcpt = pulp.LpVariable.dicts("xcpt",
                                 [(c, p, t, e) for c in C for p in P for t in T for e in E],
                                 cat='Binary')

    # Define the problem
    problem = pulp.LpProblem("Caregiver Scheduling", pulp.LpMinimize)

    # Objective function: minimize late appointments and number of caregivers per patient
    problem += (
            pulp.lpSum(xcpt[c, p, t, e] * t for c in C for p in P for t in T for e in E)  # Minimize time of appointment
    ), "Minimize_Appointment_Time_and_Caregivers_Per_Patient"

    # Constraint 1: A caregiver can care for at most one patient at a given time slot using any equipment
    for c in C:
        for t in T:
            problem += pulp.lpSum(xcpt[c, p, t, e] for p in P for e in E) <= 1, f"Caregiver_{c}time{t}"

    # Constraint 2: A patient can only attend one appointment at a given time slot
    for p in P:
        for t in T:
            problem += pulp.lpSum(xcpt[c, p, t, e] for c in C for e in E) <= 1, f"Patient_{p}time{t}"

    # Constraint 3: Ensure that caregivers are only assigned during available time slots
    constraint_counter = 0
    for c in C:
        for t in caregiver_unavailability.get(c, []):
            for p in P:
                for e in E:
                    # Add a unique counter to avoid duplicate names
                    problem += xcpt[c, p, t, e] == 0, f"Unavailable_Caregiver_{c}_{p}_{t}_{e}_{constraint_counter}"
                    constraint_counter += 1

    # Constraint 4: Ensure that patients are only assigned during available time slots
    for p in P:
        for t in patient_unavailability.get(p, []):
            for c in C:
                for e in E:
                    problem += xcpt[c, p, t, e] == 0, f"Unavailable_Patient_{p}{c}{t}_{e}"

    # Constraint 5: Equipment can only be used by one caregiver at a time
    for e in E:
        for t in T:
            problem += pulp.lpSum(xcpt[c, p, t, e] for c in C for p in P) <= 1, f"Equipment_{e}time{t}"

    # Constraint 6: Each patient must receive the specified number of appointments with the required equipment
    for p in P:
        required_equipments = dict(patient_equipment_mapping.get(p, []))
        for e, num_appointments in required_equipments.items():
            problem += pulp.lpSum(xcpt[c, p, t, e] for c in C for t in T) == num_appointments, f"Patient_{p}Equipment{e}_Appointments"

    # Constraint 7: Ensure caregivers use only equipment they are qualified for
    for c in C:
        allowed_equipments = Ce.get(c, [])
        for e in E:
            if e not in allowed_equipments:
                for p in P:
                    for t in T:
                        problem += xcpt[c, p, t, e] == 0, f"Caregiver_{c}cannot_use{e}at_time{t}for_patient{p}"

    # Constraint 8: Ensure patients are treated only with the equipment they need
    for p in P:
        required_equipments = dict(patient_equipment_mapping.get(p, []))
        for e in E:
            if e not in required_equipments:
                for c in C:
                    for t in T:
                        problem += xcpt[c, p, t, e] == 0, f"Patient_{p}doesnt_need{e}at_time{t}with_caregiver{c}"

    # Constraint 9: Ensure patients have consecutive appointments with the same caregiver
    for p in P:
        for equipment, usage_count in patient_equipment_mapping[p]:
            if usage_count > 1:
                for t in T:
                    # Add constraints for consecutive appointments
                    for c in C:
                        # Check if keys exist before adding constraints
                        keys_exist = True
                        for i in range(usage_count):
                            if (c, p, t + i, equipment) not in xcpt:
                                keys_exist = False
                                break

                        if keys_exist:
                            # Constraint to ensure that the same caregiver is assigned for consecutive appointments
                            for i in range(usage_count):
                                problem += pulp.lpSum(xcpt[c, p, t + j, equipment] for j in range(i, usage_count)) >= 1, \
                                    f"Consecutive_Appointments_Same_Caregiver_Patient_{p}_Equipment_{equipment}_Time_{t}_Caregiver_{c}_{i}"

    # Solve the problem
    status = problem.solve()

    # Prepare data for exporting to Excel
    schedule = {t: {c: "" for c in C} for t in T}  # Dictionary to store schedule (times as rows, caregivers as columns)

    if pulp.LpStatus[status] == 'Optimal':
        print("Optimal Solution Found:")
        for c in C:
            for p in P:
                for t in T:
                    for e in E:
                        if pulp.value(xcpt[c, p, t, e]) == 1:
                            schedule[t][c] = f"{p}, {e}"
                            print(f"Caregiver {c} cares for Patient {p} at time {t} with {e}")
    else:
        print("No optimal solution found. Creating a feasible schedule to maximize patient treatment.")
        # Define a new problem to maximize the number of treated patients
        problem_feasible = pulp.LpProblem("Feasible_Caregiver_Scheduling", pulp.LpMaximize)

        # New objective function: maximize the number of patients treated
        problem_feasible += pulp.lpSum(xcpt[c, p, t, e] for c in C for p in P for t in T for e in E), "Maximize_Patients_Treated"

        # Apply the same constraints but ignore constraints 6 & 9
        for c in C:
            for t in T:
                problem_feasible += pulp.lpSum(xcpt[c, p, t, e] for p in P for e in E) <= 1, f"Caregiver_{c}time{t}"

        for p in P:
            for t in T:
                problem_feasible += pulp.lpSum(xcpt[c, p, t, e] for c in C for e in E) <= 1, f"Patient_{p}time{t}"

        for c in C:
            for t in caregiver_unavailability.get(c, []):
                for p in P:
                    for e in E:
                        problem_feasible += xcpt[c, p, t, e] == 0, f"Unavailable_Caregiver_{c}{p}{t}_{e}"

        for p in P:
            for t in patient_unavailability.get(p, []):
                for c in C:
                    for e in E:
                        problem_feasible += xcpt[c, p, t, e] == 0, f"Unavailable_Patient_{p}{c}{t}_{e}"

        for e in E:
            for t in T:
                problem_feasible += pulp.lpSum(xcpt[c, p, t, e] for c in C for p in P) <= 1, f"Equipment_{e}time{t}"

        for c in C:
            allowed_equipments = Ce.get(c, [])
            for e in E:
                if e not in allowed_equipments:
                    for p in P:
                        for t in T:
                            problem_feasible += xcpt[c, p, t, e] == 0, f"Caregiver_{c}cannot_use{e}at_time{t}for_patient{p}"

        for p in P:
            required_equipments = dict(patient_equipment_mapping.get(p, []))
            for e in E:
                if e not in required_equipments:
                    for c in C:
                        for t in T:
                            problem_feasible += xcpt[c, p, t, e] == 0, f"Patient_{p}doesnt_need{e}at_time{t}with_caregiver{c}"

        for p in P:
            required_equipments = dict(patient_equipment_mapping.get(p, []))
            for e, num_appointments in required_equipments.items():
                problem_feasible += pulp.lpSum(
                    xcpt[c, p, t, e] for c in C for t in T) <= num_appointments, f"Patient_{p}Equipment{e}_Appointments"

        # Solve the modified problem
        status_feasible = problem_feasible.solve()

        if pulp.LpStatus[status_feasible] == 'Optimal':
            for c in C:
                for p in P:
                    for t in T:
                        for e in E:
                            if pulp.value(xcpt[c, p, t, e]) == 1:
                                schedule[t][c] = f"{p}, {e}"
        else:
            print("No feasible solution found. Check constraints and availability.")

    save_schedule_to_excel(schedule, input_file_path, output_file_path)


def save_schedule_to_excel(schedule, input_file, output_file):
    # Define sets
    C, P, E, Ce, caregiver_unavailability, patient_unavailability, patient_equipment_mapping = \
        excel_sheets_to_items(input_file)

    # Convert schedule to DataFrame and transpose
    df_schedule = pd.DataFrame(schedule).transpose()
    df_schedule.index = [f"{hour}:00" for hour in T]  # Set the time slots (rows) from 8:00 to 17:00
    df_schedule.columns = C

    # Ensure the output directory exists
    output_dir = os.path.dirname(output_file)
    os.makedirs(output_dir, exist_ok=True)

    # Export the DataFrame to an Excel file
    df_schedule.to_excel(output_file, index=True, engine='openpyxl')

    # Load the workbook and select the active worksheet
    workbook = load_workbook(output_file)
    worksheet = workbook.active

    # Equipment colors mapping
    equipment_colors = {equipment: generate_random_color() for equipment in E}

    # Apply colors to cells based on equipment
    for row_idx, row in enumerate(
            worksheet.iter_rows(min_row=2, min_col=2, max_row=worksheet.max_row, max_col=worksheet.max_column),
            start=8):  # starting at 8:00 (adjust based on time slots)
        for col_idx, cell in enumerate(row, start=0):
            caregiver_name = C[col_idx]

            # Check caregiver unavailability
            if caregiver_name in caregiver_unavailability and row_idx in caregiver_unavailability[caregiver_name]:
                cell.value = "Unavailable"
            else:
                # Extract patient and equipment from the cell value
                cell_value = cell.value
                if cell_value:
                    patient, equipment = cell_value.split(", ")
                    # Apply color based on equipment
                    if equipment in equipment_colors:
                        aRGB_color = equipment_colors[equipment]
                        cell.fill = PatternFill(start_color=aRGB_color, end_color=aRGB_color, fill_type="solid")

    # Add a map of equipment and colors to the Excel file
    equipment_color_map_start_row = worksheet.max_row + 3
    worksheet.cell(row=equipment_color_map_start_row, column=1, value="Equipment")
    worksheet.cell(row=equipment_color_map_start_row, column=2, value="Color")

    for idx, (equipment, color) in enumerate(equipment_colors.items(), start=equipment_color_map_start_row + 1):
        worksheet.cell(row=idx, column=1, value=equipment)
        color_cell = worksheet.cell(row=idx, column=2)
        color_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        # Clear the text in the color cell
        color_cell.value = ""

    # Now, create a new sheet for unscheduled patients
    unscheduled_patients = create_a_list_of_patients_who_are_not_part_of_the_schedule(schedule,
                                                                                      patient_equipment_mapping)

    # Create a new worksheet for unscheduled patients
    unscheduled_worksheet = workbook.create_sheet("Unscheduled Patients")
    unscheduled_worksheet.cell(row=1, column=1, value="Patient")
    unscheduled_worksheet.cell(row=1, column=2, value="Equipment")
    unscheduled_worksheet.cell(row=1, column=3, value="Count")

    # Write unscheduled patients data to the new sheet
    row_idx = 2
    for patient, equipment_list in unscheduled_patients.items():
        for equipment, count in equipment_list:
            unscheduled_worksheet.cell(row=row_idx, column=1, value=patient)  # Write the patient's name
            unscheduled_worksheet.cell(row=row_idx, column=2, value=equipment)  # Write the equipment
            unscheduled_worksheet.cell(row=row_idx, column=3, value=count)  # Write the count
            row_idx += 1

    # Save the workbook with the new "Unscheduled Patients" sheet
    workbook.save(output_file)
    print(
        f"Solution exported to {output_file} with colored cells, equipment color map, and unscheduled patients list in a new sheet")


def generate_random_color():
    alpha = 'FF'  # Fully opaque
    red = format(random.randint(150, 255), '02X')
    green = format(random.randint(150, 255), '02X')
    blue = format(random.randint(150, 255), '02X')
    return f"{alpha}{red}{green}{blue}"

def create_a_list_of_patients_who_are_not_part_of_the_schedule(schedule, patient_equipment_mapping):
    # Copy the original patient-equipment mapping to avoid modifying it
    patients_unscheduled = copy.deepcopy(patient_equipment_mapping)

    # Extract the patients and equipment from the schedule
    scheduled_patients_and_equipments = extract_patients_from_schedule(schedule)

    # Build a dictionary of scheduled patients and their equipment counts
    patients_and_equipments_dict = {}
    for patient, equipment in scheduled_patients_and_equipments:
        if patient not in patients_and_equipments_dict:
            patients_and_equipments_dict[patient] = {}
        if equipment in patients_and_equipments_dict[patient]:
            patients_and_equipments_dict[patient][equipment] += 1
        else:
            patients_and_equipments_dict[patient][equipment] = 1

    # Update the unscheduled patients
    updated_patients_unscheduled = {}
    for patient, equipment_list in patients_unscheduled.items():
        updated_equipment_list = []
        for equipment, count in equipment_list:
            # Check if the equipment is already scheduled and reduce its count
            if patient in patients_and_equipments_dict and equipment in patients_and_equipments_dict[patient]:
                remaining_count = count - patients_and_equipments_dict[patient][equipment]
            else:
                remaining_count = count  # No conflict if the equipment isn't scheduled

            # Add equipment to the updated list if there's any remaining count
            if remaining_count > 0:
                updated_equipment_list.append((equipment, remaining_count))

        # Add the patient back to the unscheduled list if there's remaining equipment
        if updated_equipment_list:
            updated_patients_unscheduled[patient] = updated_equipment_list

    return updated_patients_unscheduled

def extract_patients_from_schedule(schedule):
    patient_list = []
    for hour, caregivers in schedule.items():
        for caregiver, patient_equipment in caregivers.items():
            if patient_equipment:  # Check if there's a patient assigned
                if patient_equipment.strip().lower() == "unavailable":  # Skip "Unavailable" entries
                    continue
                try:
                    patient, equipment = patient_equipment.split(', ')
                    patient_list.append((patient.strip(), equipment.strip()))
                except ValueError as e:
                    print(f"Error splitting patient_equipment '{patient_equipment}': {e}")
    return patient_list


def main():
    input_file = r"C:\Users\morsh\Desktop\personal_projects\soroka_solution\small_rehabilitation_data.xlsx"
    output_file = r"C:\Users\morsh\Desktop\personal_projects\soroka_solution\generated_schedule.xlsx"
    create_original_schedule(input_file, output_file)

if __name__ == "__main__":
    main()